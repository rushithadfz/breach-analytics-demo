"""The exposure table API surface: list/filter persons, drill into a
person's flags with source-document evidence (brief section 2/4)."""
from __future__ import annotations

import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.api.v1.deps import require_api_key
from app.db.base import get_db
from app.db.models import (
    Document, EntityLink, Extraction, ExtractionMethod, ExposureFlag, FlagEvidence, PiiCategory,
    Person, ReviewStatus,
)
from app.schemas.api import (
    EvidenceOut, ExposureFlagOut, ExposureSummary, PersonDetail, PersonListItem,
)

router = APIRouter(prefix="/persons", tags=["persons"], dependencies=[Depends(require_api_key)])


@router.get("", response_model=list[PersonListItem])
def list_persons(
    category: PiiCategory | None = None,
    search: str | None = Query(default=None, description="substring match on name"),
    limit: int = Query(default=50, le=500),
    offset: int = 0,
    db: Session = Depends(get_db),
):
    stmt = select(Person).options(selectinload(Person.flags))
    if search:
        stmt = stmt.where(Person.best_known_full_name.ilike(f"%{search}%"))
    if category is not None:
        # Filter in SQL, before LIMIT. This used to be a Python filter
        # applied to the page after it was fetched, which silently meant
        # "matches among the first 100 people" rather than "the first 100
        # matches" — clicking a chart bar reading 206 landed on a table
        # of 99. EXISTS rather than a join so a person with several flags
        # in one category still yields one row.
        stmt = stmt.where(Person.flags.any(ExposureFlag.category == category))

    stmt = stmt.order_by(Person.id).offset(offset).limit(limit)
    people = db.execute(stmt).scalars().all()

    return [
        PersonListItem(
            id=p.id, person_uid=p.person_uid, best_known_full_name=p.best_known_full_name,
            dob=p.dob, review_status=p.review_status.value,
            flag_categories=[f.category.value for f in p.flags],
        )
        for p in people
    ]


@router.get("/summary", response_model=ExposureSummary)
def exposure_summary(db: Session = Depends(get_db)):
    """Category and review-status counts for the dashboard. Declared
    before /{person_id} so the literal path wins the match."""
    by_category = {
        cat.value: n for cat, n in db.execute(
            select(ExposureFlag.category, func.count()).group_by(ExposureFlag.category)
        ).all()
    }
    by_review_status = {
        st.value: n for st, n in db.execute(
            select(Person.review_status, func.count()).group_by(Person.review_status)
        ).all()
    }
    total_persons = db.execute(select(func.count()).select_from(Person)).scalar_one()
    needing_review = db.execute(
        select(func.count(func.distinct(ExposureFlag.person_id)))
        .where(ExposureFlag.review_status == ReviewStatus.needs_review)
    ).scalar_one()

    return ExposureSummary(
        total_persons=total_persons,
        persons_needing_review=needing_review,
        total_flags=sum(by_category.values()),
        by_category=by_category,
        by_review_status=by_review_status,
    )


@router.get("/export.csv")
def export_exposure_table_csv(db: Session = Depends(get_db)):
    """Same content as sheet 1 of the XLSX, for tooling that wants text.

    This originally carried only identity and a yes/blank per category —
    one of the four column groups brief §2 specifies, and not the one
    that makes a row defensible. The per-flag evidence references live in
    the XLSX's second sheet; CSV is a single table by definition, so it
    carries the source-document count and points at the richer export.
    """
    categories = [c.value for c in PiiCategory if c != PiiCategory.full_name]
    people, aliases, sources, resolution_confidence = _exposure_rows(db)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "person_uid", "best_known_full_name", "known_aliases", "dob",
        "review_status", "resolution_confidence", "source_documents",
        *[f"{c}_evidence_count" for c in categories],
    ])
    for p in people:
        flags = {f.category.value: f for f in p.flags}
        all_refs = set()
        for cat in categories:
            all_refs |= sources.get((p.id, cat), set())
        variants = sorted(aliases.get(p.id, set()) - {p.best_known_full_name})
        conf = resolution_confidence.get(p.id)
        writer.writerow([
            p.person_uid, p.best_known_full_name, "; ".join(variants), p.dob or "",
            p.review_status.value,
            round(conf, 3) if conf is not None else "",
            len({r.split("#")[0] for r in all_refs}),
            *[len(sources.get((p.id, c), set())) if c in flags else "" for c in categories],
        ])

    buf.seek(0)
    return StreamingResponse(buf, media_type="text/csv",
                              headers={"Content-Disposition": "attachment; filename=exposure_table.csv"})


def _exposure_rows(db: Session):
    """The exposure table with everything brief §2 asks a row to carry.

    The CSV export predates this and answers only "is this category
    exposed" — which is the boolean column group and none of the other
    three. §2 also requires known aliases, a count of source documents,
    document references per flag, and confidence alongside review status.
    A row that cannot show its sources is, in the brief's own words,
    worthless in front of a regulator.

    Gathered in three bulk queries rather than per person, because the
    obvious per-person version issues 3N round trips and this endpoint
    runs over the whole table.
    """
    people = db.execute(
        select(Person).options(selectinload(Person.flags)).order_by(Person.id)
    ).scalars().all()

    aliases: dict[int, set[str]] = {}
    for person_id, value in db.execute(
        select(EntityLink.person_id, Extraction.normalized_value)
        .join(Extraction, Extraction.id == EntityLink.extraction_id)
        .where(Extraction.category == PiiCategory.full_name)
    ).all():
        if value:
            aliases.setdefault(person_id, set()).add(value)

    # Source documents per (person, category), for the "document
    # references per flag" column. Distinct because one document can
    # yield several extractions of the same category.
    sources: dict[tuple[int, str], set[str]] = {}
    for person_id, category, relpath, page in db.execute(
        select(EntityLink.person_id, Extraction.category, Document.relpath, Extraction.page_number)
        .join(Extraction, Extraction.id == EntityLink.extraction_id)
        .join(Document, Document.id == Extraction.document_id)
    ).all():
        ref = f"{relpath}#p{page}" if page else relpath
        sources.setdefault((person_id, category.value), set()).add(ref)

    resolution_confidence: dict[int, float] = {}
    for person_id, worst in db.execute(
        select(EntityLink.person_id, func.min(EntityLink.confidence)).group_by(EntityLink.person_id)
    ).all():
        resolution_confidence[person_id] = worst

    return people, aliases, sources, resolution_confidence


@router.get("/export.xlsx")
def export_exposure_table_xlsx(db: Session = Depends(get_db)):
    """Two sheets, because one row per person cannot also hold one row
    per piece of evidence without becoming unreadable.

    Sheet 1 is the denormalized table legal consumes. Sheet 2 is the
    defensibility layer: every flag with its confidence and the exact
    documents and pages behind it, so a claim can be checked without
    opening the app.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    categories = [c.value for c in PiiCategory if c != PiiCategory.full_name]
    people, aliases, sources, resolution_confidence = _exposure_rows(db)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A19")

    def write_header(ws, labels):
        ws.append(labels)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Exposure table"
    write_header(ws, [
        "Person ID", "Best-known name", "Known aliases", "Date of birth",
        "Review status", "Resolution confidence", "Source documents",
        *[c.replace("_", " ").title() for c in categories],
    ])

    for p in people:
        flags = {f.category.value: f for f in p.flags}
        all_refs = set()
        for cat in categories:
            all_refs |= sources.get((p.id, cat), set())
        variants = sorted(aliases.get(p.id, set()) - {p.best_known_full_name})
        conf = resolution_confidence.get(p.id)
        ws.append([
            p.person_uid,
            p.best_known_full_name,
            "; ".join(variants),
            p.dob or "",
            p.review_status.value,
            round(conf, 3) if conf is not None else "",
            len({r.split("#")[0] for r in all_refs}),
            # The count, not "yes": it is the boolean (non-empty means
            # exposed) and the evidence count in one column, which keeps
            # the table narrow enough to read.
            *[len(sources.get((p.id, c), set())) if c in flags else "" for c in categories],
        ])

    widths = [12, 26, 30, 13, 16, 20, 17] + [15] * len(categories)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    detail = wb.create_sheet("Flag evidence")
    write_header(detail, [
        "Person ID", "Best-known name", "Category", "Confidence",
        "Review status", "Source documents (path#page)",
    ])
    for p in people:
        for flag in sorted(p.flags, key=lambda f: f.category.value):
            refs = sorted(sources.get((p.id, flag.category.value), set()))
            detail.append([
                p.person_uid, p.best_known_full_name, flag.category.value,
                round(flag.confidence, 3), flag.review_status.value,
                "; ".join(refs),
            ])
    for i, w in enumerate([12, 26, 20, 12, 16, 90], start=1):
        detail.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exposure_table.xlsx"},
    )


@router.get("/{person_id}", response_model=PersonDetail)
def get_person(person_id: int, db: Session = Depends(get_db)):
    person = db.execute(
        select(Person).options(selectinload(Person.flags)).where(Person.id == person_id)
    ).scalar_one_or_none()
    if person is None:
        raise HTTPException(status_code=404, detail="person not found")
    flags = [ExposureFlagOut.model_validate(f) for f in person.flags]
    return PersonDetail(id=person.id, person_uid=person.person_uid, best_known_full_name=person.best_known_full_name,
                         dob=person.dob, review_status=person.review_status.value, flags=flags)


@router.get("/{person_id}/flags/{flag_id}/evidence", response_model=list[EvidenceOut])
def get_flag_evidence(person_id: int, flag_id: int, db: Session = Depends(get_db)):
    flag = db.get(ExposureFlag, flag_id)
    if flag is None or flag.person_id != person_id:
        raise HTTPException(status_code=404, detail="flag not found for this person")

    rows = db.execute(
        select(FlagEvidence, Extraction, Document)
        .join(Extraction, FlagEvidence.extraction_id == Extraction.id)
        .join(Document, Extraction.document_id == Document.id)
        .where(FlagEvidence.exposure_flag_id == flag_id)
    ).all()
    llm_methods = {ExtractionMethod.llm_cheap, ExtractionMethod.llm_strong}
    return [
        EvidenceOut(
            extraction_id=ext.id, document_id=doc.id, document_relpath=doc.relpath,
            document_type=doc.sniffed_type.value, category=ext.category.value,
            passage=ext.passage, confidence=ext.confidence, method=ext.method.value,
            page_number=ext.page_number, record_key=ext.record_key,
            page_is_approximate=ext.method in llm_methods,
        )
        for _, ext, doc in rows
    ]
